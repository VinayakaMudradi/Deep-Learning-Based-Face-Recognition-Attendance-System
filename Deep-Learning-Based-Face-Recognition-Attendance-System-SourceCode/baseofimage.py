# -*- coding: utf-8 -*-
"""
Created on Tue Mar 18 12:34:50 2020

@author: VINAYAKA MUDRADI
"""
# code for building the data base of our model
# we prepare 5 photos of everyone, change them and apply LBP and face detection

import os
from os import listdir
from os.path import isfile, join

import cv2
import numpy
import numpy as np
from numpy import *
from random import randint
import MySQLdb as msd
from openpyxl import Workbook
from openpyxl.styles import Font

book=Workbook()
sheet=book.active
sheet.cell(row=1, column=1, value="USN").font=Font(bold=True)
sheet.cell(row=1, column=2, value="NAME").font=Font(bold=True)
sheet.cell(row=1, column=3, value="EMAIL").font=Font(bold=True)
sheet.cell(row=1, column=4, value="TOTAL CLASSES").font=Font(bold=True)
sheet.column_dimensions['D'].width=15
    #sheet.column_dimensions['D'].width=15
sheet.cell(row=1, column=5, value="CLASSES ATTENDED").font=Font(bold=True)
sheet.column_dimensions['E'].width=19
sheet.cell(row=1, column=6, value="PERCENTAGE(%)").font=Font(bold=True)
sheet.column_dimensions['F'].width=15



try:
    db= msd.connect('localhost','root','vinayaka','attendancesystem')
    print("\n [INFO] Database Connection Successful..!\n Connected to LocalHost")
    
except msd.Error as e:
    print("\n [INFO] We got this error Specified below: "+e)
    exit(0)
# import the photo of someone and change randomly for the base of images.
# here i take 5 photos for everyone, change them into 100 photos for each photo
# so totally 500 photos for everyone, totally 2000 photos as the input of CNN

def larger(img):
    # get the number of rows and cols of picture
    rows, cols = img.shape[:2]
    # take a random number to use
    num = randint(0, 24)

    pts1 = np.float32([[num, num], [cols - num, num], [num, rows - num], [cols - num, rows - num]])
    pts2 = np.float32([[0, 0], [cols, 0], [0, rows], [cols, rows]])
    M = cv2.getPerspectiveTransform(pts1, pts2)
    dst = cv2.warpPerspective(img, M, (cols, rows))
    return dst

def smaller(img):
    # get the number of rows and cols of picture
    rows, cols = img.shape[:2]
    # take a random number to use
    num = randint(0, 24)

    pts1 = np.float32([[num, num], [cols - num, num], [num, rows - num], [cols - num, rows - num]])
    pts2 = np.float32([[0, 0], [cols, 0], [0, rows], [cols, rows]])
    M = cv2.getPerspectiveTransform(pts2, pts1)
    dst = cv2.warpPerspective(img, M, (cols, rows))
    return dst

def lighter(img):
    # copy the basic picture, avoid the change of the basic one
    dst = cv2.copyMakeBorder(img, 0, 0, 0, 0, cv2.BORDER_REPLICATE)
    # get the number of rows and cols of picture
    rows, cols = dst.shape[:2]
    # take a random number to use
    num = randint(20,50)

    for xi in range(0, cols):
        for xj in range(0, rows):
            for i in range(0, 3):
                if dst[xj, xi, i] <= 255 - num:
                    dst[xj, xi, i] = int(dst[xj, xi, i] + num)
                else:
                    dst[xj, xi, i] = 255
    return dst

def darker(img):
    # copy the basic picture, avoid the change of the basic one
    dst = cv2.copyMakeBorder(img, 0, 0, 0, 0, cv2.BORDER_REPLICATE)
    # get the number of rows and cols of picture
    rows, cols = img.shape[:2]
    # take a random number to use
    num = randint(20,50)

    for xi in range(0, cols):
        for xj in range(0, rows):
            for i in range(0, 3):
                if dst[xj, xi, i] >= num:
                    dst[xj, xi, i] = int(dst[xj, xi, i] - num)
                else:
                    dst[xj, xi, i] = 0
    return dst

def moveright(img):
    # get the number of rows and cols of picture
    rows,cols = img.shape[:2]
    # take a random number to use
    num = randint(1, 2)

    M = np.float32([[1,0,num],[0,1,0]])
    dst = cv2.warpAffine(img,M,(cols,rows))
    return dst

def moveleft(img):
    # get the number of rows and cols of picture
    rows,cols = img.shape[:2]
    # take a random number to use
    num = randint(1, 2)

    M = np.float32([[1,0,-num],[0,1,0]])
    dst = cv2.warpAffine(img,M,(cols,rows))
    return dst

def movetop(img):
    # get the number of rows and cols of picture
    rows,cols = img.shape[:2]
    # take a random number to use
    num = randint(1, 2)

    M = np.float32([[1,0,0],[0,1,-num]])
    dst = cv2.warpAffine(img,M,(cols,rows))
    return dst

def movebot(img):
    # get the number of rows and cols of picture
    rows,cols = img.shape[:2]
    # take a random number to use
    num = randint(1, 2)

    M = np.float32([[1,0,0],[0,1,num]])
    dst = cv2.warpAffine(img,M,(cols,rows))
    return dst

def turnright(img):
    # get the number of rows and cols of picture
    rows, cols = img.shape[:2]
    # take a random number to use
    num = randint(3,6)

    M = cv2.getRotationMatrix2D((cols / 2, rows / 2), -num, 1)
    dst = cv2.warpAffine(img, M, (cols, rows))
    return dst

def turnleft(img):
    # get the number of rows and cols of picture
    rows, cols = img.shape[:2]
    # take a random number to use
    num = randint(3,6)

    M = cv2.getRotationMatrix2D((cols / 2, rows / 2), num, 1)
    dst = cv2.warpAffine(img, M, (cols, rows))
    return dst

def changeandsave(name,time,choice,img,i):
    # the new name of changed picture is "changed?.png",? means it's the ?-th picture changed
    #print("came here change and save")
    name = 'changedphoto/' + str(name) + '/' +str(time) + '_changed' + str(i) + '.jpg'
    # do different changes by the choice
    if choice == 1:
        newimg = larger(img)
    elif choice == 2:
        newimg = smaller(img)
    elif choice == 3:
        newimg = lighter(img)
    elif choice == 4:
        newimg = darker(img)
    elif choice == 5:
        newimg = moveright(img)
    elif choice == 6:
        newimg = moveleft(img)
    elif choice == 7:
        newimg = movetop(img)
    elif choice == 8:
        newimg = movebot(img)
    elif choice == 9:
        newimg = turnleft(img)
    elif choice == 10:
        newimg = turnright(img)
    # save the new picture
    cv2.imwrite(name, newimg)



train_images=os.listdir('images')
#count=0
studentcount=0
row_count=1
for train_image in train_images:
    print("\n [INFO] Processing images of "+str(train_image))
    #count=count+1
    studentcount=studentcount+1
    row_count=row_count+1
    cursor=db.cursor()
    Query1="UPDATE student SET ID='%d' WHERE Name='%s'" %(int(studentcount), str(train_image))
    cursor.execute(Query1)
    db.commit()
    Query2="SELECT * FROM student WHERE Name='%s'"%(str(train_image))
    cursor.execute(Query2)
    student_info=cursor.fetchall()
    sheet.cell(row=row_count, column=1, value=student_info[0][0])
    sheet.column_dimensions['A'].width=15
    sheet.cell(row=row_count, column=2, value=student_info[0][1])
    sheet.column_dimensions['B'].width=25
    sheet.cell(row=row_count, column=3, value=student_info[0][2])
    sheet.column_dimensions['C'].width=35
    
    #sheet.cell(row=row_count, column=4, value="")
    
    os.mkdir('changedphoto/'+train_image)
    images=os.listdir(os.path.join('images',train_image))
    count=0
    for image in images:
        count=count+1
        img=cv2.imread('images'+'/'+str(train_image)+'/'+str(image),1)
        for i in range(1,101):
            choice = randint(1,10)
            changeandsave(train_image,count,choice,img,i)
# take fu's 5 photos, change each photo into 100 photos, so totally 500
book.save('Attendance_sheet.xlsx')
print("\n [INFO] Attendance sheet is ready.")



# and we need to use opencv to apply face detection and LPB of every image
# here we have totally 2000 photos, 500 for everyone

# the file of OPENCV for face detection
#path( = '/home/pi/Desktop/opencv-2.4.10/data/haarcascades/'


# 2 fonctions for LBP
def thresholded(center, pixels):
    out = []
    for a in pixels:
        if a >= center:
            out.append(1)
        else:
            out.append(0)
    return out


def get_pixel_else_0(l, idx, idy, default=0):
    try:
        return l[idx, idy]
    except IndexError:
        return default

#print("cascade")
#path = '/home/user/opencv-3.1.0/data/haarcascades/'
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
fcount=0
#print("came here")
changed_folders=os.listdir('changedphoto')
for changed_folder in changed_folders:
    print("\n [INFO] Face detection and LBP extraction of images of "+str(changed_folder))
    fcount=fcount+1
    #print("hello")
    mypath=os.path.join('changedphoto',changed_folder)
    print(" "+mypath)
    onlyfiles=[f for f in listdir(mypath) if isfile(join(mypath, f))]
    #print(onlyfiles)
    images= numpy.empty(len(onlyfiles),dtype=object)
    for n in range(0,len(onlyfiles)):
        images[n]=cv2.imread(join(mypath, onlyfiles[n]))
        newgray=cv2.cvtColor(images[n],cv2.COLOR_BGR2GRAY)
        faces=face_cascade.detectMultiScale(newgray,scaleFactor=1.1,minNeighbors=5,minSize=(30,30),flags=cv2.CASCADE_SCALE_IMAGE)
        for (x,y,w,h) in faces:
            x=x
        cropped=newgray[y:y+h,x:x+w]
        result=cv2.resize(cropped,(48,48),interpolation=cv2.INTER_LINEAR)
        transformed_img=cv2.copyMakeBorder(result,0,0,0,0,cv2.BORDER_REPLICATE)
        for x in range(0,len(result)):
            for y in range(0,len(result[0])):
                center=result[x,y]
                top_left=get_pixel_else_0(result, x-1, y-1)
                top_up=get_pixel_else_0(result,x,y-1)
                top_right=get_pixel_else_0(result,x+1,y-1)
                right=get_pixel_else_0(result,x+1,y)
                left=get_pixel_else_0(result,x-1,y)
                bottom_left=get_pixel_else_0(result,x-1,y+1)
                bottom_right=get_pixel_else_0(result,x+1,y+1)
                bottom_down=get_pixel_else_0(result,x,y+1)
                values=thresholded(center, [top_left,top_up,top_right,right,bottom_right,bottom_down,bottom_left,left])
                weights=[1,2,4,8,16,32,64,128]
                res=0
                for a in range(0,len(values)):
                    res+=weights[a]*values[a]
                transformed_img.itemset((x,y),res)
        lbp = transformed_img[1:47, 1:47]
        name = 'CNNdata/'+str(fcount)+'_'+str(changed_folder)+'_' + str(n) + '.jpg'
        cv2.imwrite(name, lbp)
                
# we apply LBP and face detection to all photos and save the new photos
print("\n [INFO] Face data is ready to train the CNN model.")
wait_key=input()