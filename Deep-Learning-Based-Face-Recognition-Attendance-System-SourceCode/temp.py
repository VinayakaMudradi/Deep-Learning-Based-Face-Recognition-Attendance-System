# -*- coding: utf-8 -*-
"""
Created on Tue Apr 14 21:45:08 2020

@author: VINAYAKA MUDRADI
"""


# code for testing the model by some existing normal photos

from keras.models import Sequential
from keras.models import model_from_json
from keras.utils import np_utils
import MySQLdb as msd
import cv2
from openpyxl.styles import Font
from openpyxl.styles import colors
import numpy as np
import datetime
from skimage import io
from sklearn.model_selection import train_test_split
import openpyxl
import os
from os import listdir
from os.path import isfile, join



book=openpyxl.load_workbook('Attendance_sheet.xlsx')
sheet=book.active
max_col_number=sheet.max_column

try:
    db= msd.connect('localhost','root','vinayaka','attendancesystem')
    print("\n [INFO] Database Connection Successful..!\n Connected to LocalHost")
    
except msd.Error as e:
    print("\n [INFO] We got this error Specified below: "+e)
    exit(0)
    
# 2 fonctions for LBP
def thresholded(center, pixels):
    out = []
    for a in pixels:
        if a >= center:
            out.append(1)
        else:
            out.append(0)
    return out


def get_pixel_else_0(l, idx, idy, default=0):
    try:
        return l[idx, idy]
    except IndexError:
        return default

# the file of OPENCV for face detection
#path = '/home/pi/Desktop/opencv-2.4.10/data/haarcascades/'
#path = '/home/user/opencv-3.1.0/data/haarcascades/'
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

person_name=[]
names=os.listdir('changedphoto')
for name in names:
    person_name.append(name)
    

# load json and create model
print("\n [INFO] Loading the trained CNN model.")
json_file = open('model.json', 'r')
loaded_model_json = json_file.read()
json_file.close()
loaded_model = model_from_json(loaded_model_json)
# load weights into new model
loaded_model.load_weights("model.h5")
#print("Loaded model from disk")
DatasetPath = []
for i in os.listdir('photooftest'):
    DatasetPath.append(os.path.join('photooftest', i))

imageData = []
imageName = []

# then read the photos, find the face in the photo
# crop the part of face, apply LBP and resize into 46*46
# then save the new photos and filenames
fname=[]
print("\n [INFO] Images in the folder for prediction.")
c=0
for i in DatasetPath:
    print(i)
    imgRead = cv2.imread(i,0) # read the photo by gray
    imageName.append(str(i))
    faces = face_cascade.detectMultiScale(
        imgRead,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(30, 30),
        flags=cv2.CASCADE_SCALE_IMAGE #OPENCV version 3.x
        #flags = cv2.cv.CV_HAAR_SCALE_IMAGE #OPENCV version 2.x
    )
    #c=0
    for (x, y, w, h) in faces:
        x=x

        cropped = imgRead[y:y + h, x:x + w]
        c=c+1
        cv2.imwrite('face/'+str(c)+'.jpg',cropped)
        result = cv2.resize(cropped, (48, 48), interpolation=cv2.INTER_LINEAR)  # OPENCV 3.x
    
        transformed_img = cv2.copyMakeBorder(result, 0, 0, 0, 0, cv2.BORDER_REPLICATE)

        for x in range(0, len(result)):
            for y in range(0, len(result[0])):
                center = result[x, y]
                top_left = get_pixel_else_0(result, x - 1, y - 1)
                top_up = get_pixel_else_0(result, x, y - 1)
                top_right = get_pixel_else_0(result, x + 1, y - 1)
                right = get_pixel_else_0(result, x + 1, y)
                left = get_pixel_else_0(result, x - 1, y)
                bottom_left = get_pixel_else_0(result, x - 1, y + 1)
                bottom_right = get_pixel_else_0(result, x + 1, y + 1)
                bottom_down = get_pixel_else_0(result, x, y + 1)

                values = thresholded(center, [top_left, top_up, top_right, right, bottom_right,
                                          bottom_down, bottom_left, left])

                weights = [1, 2, 4, 8, 16, 32, 64, 128]
                res = 0
                for a in range(0, len(values)):
                    res += weights[a] * values[a]

                transformed_img.itemset((x, y), res)

    # we only use the part (1,1) to (46,46) of the result img.
    # original img: 0-47, after resize: 1-46
        lbp = transformed_img[1:47, 1:47]  # here 1 included, 47 not included

        imageData.append(lbp)


def col_string(n):
    string=""
    while n>0:
        n,remainder=divmod(n-1,26)
        string=chr(65+remainder)+string
    return string

dimentionChar=col_string(max_col_number+1)
sheet.column_dimensions[dimentionChar].width=18

sheet.cell(row=1,column=max_col_number+1,value=datetime.datetime.now()).font=Font(bold=True)
cursor=db.cursor()
cursor.execute("SELECT COUNT(*) FROM student")
number_of_students=cursor.fetchall()
#print("Number:"+str(number_of_students))
present_students=[]
for i in range(0,len(imageData)):
    person_found=0
    c=np.array(imageData[i])
    c=np.array(c)
    c=c.reshape(1,46,46,1)
    c=c.astype('float32')
    c/=255
    predictions=loaded_model.predict(c)
    print(predictions)
    index_val=np.argmax(predictions,axis=1)
    print(index_val)
    prob=predictions[0][index_val]
    
    
    if prob>0.90:
            person_found=1
            Query="SELECT * FROM student WHERE ID='%d'" %(int(index_val+1))
            cursor.execute(Query)
            result=cursor.fetchall()
            
            print(" [INFO] USN:"+str(result[0][0])+"\n"+" Name:"+str(result[0][1]))
            print(person_name[index_val[0]])
            print("\n")
            sheet.cell(row=int(index_val+2),column=max_col_number+1,value="Present").font=Font(color='228B22')
            present_students.append(index_val)
    if person_found==0:
        print(" [INFO] Not match")
        print("\n")
for j in range(number_of_students[0][0]):
    if j in present_students:
        continue
    sheet.cell(row=int(j+2),column=max_col_number+1,value="Absent").font=Font(color=colors.RED)

for k in range(number_of_students[0][0]):
    present_count=0
    total_count=0
    for l in range(7, max_col_number+2):
        total_count=total_count+1
        if sheet.cell(row=k+2,column=l).value =='Present':
            present_count=present_count+1
    percentage=(present_count/total_count)*100
    percentage=float("{:.2f}".format(percentage))
    sheet.cell(row=int(k+2),column=5,value=present_count).font=Font(color=colors.BLUE)
    sheet.cell(row=int(k+2),column=4,value=total_count).font=Font(color=colors.BLUE)
    sheet.cell(row=int(k+2),column=6,value=percentage).font=Font(color=colors.BLUE)
    
book.save('Attendance_sheet.xlsx')
print("\n [INFO] Attendance status updated in Excel sheet.")
print("\n [INFO] Prediction completed successfully.")
wait_key=input()
            
